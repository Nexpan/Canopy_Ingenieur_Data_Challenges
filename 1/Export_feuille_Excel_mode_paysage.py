import argparse
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
#from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
#from PyPDF2 import PdfReader, PdfWriter


def get_excel_sheet_dimensions(excel_file, sheet_name):
    # Charger le classeur Excel
    workbook = openpyxl.load_workbook(excel_file)

    # Vérifier si l'onglet existe
    if sheet_name not in workbook.sheetnames:
        print(f"La feuille '{sheet_name}' n'existe pas dans le fichier Excel.")
        return None

    # Sélectionner l'onglet spécifié
    sheet = workbook[sheet_name]

    # Obtenir les dimensions de la feuille
    min_row = sheet.min_row
    max_row = sheet.max_row
    min_col = sheet.min_column
    max_col = sheet.max_column 

    return min_row, max_row, min_col, max_col

def excel_sheet_to_pdf(excel_file, sheet_name, output_pdf):
    # Charger le classeur Excel
    workbook = openpyxl.load_workbook(excel_file)

    # Sélectionner la feuille spécifiée
    sheet = workbook[sheet_name]

    # Obtenir les dimensions de la page
    width, height = A4

    # Créer un PDF
    c = canvas.Canvas(output_pdf, pagesize=(landscape(A4)))

    # Définir les marges
    right_margin = -170
    top_margin = -250

    # Commencer l'impression des cellules dans le PDF
    for row in sheet.iter_rows():
        for cell in row:
            cell_value = str(cell.value) 
            x = cell.column * 200 # Décaler horizontalement pour éviter les collisions
            y = height - cell.row * 18.5  # Décaler verticalement pour éviter les collisions
            c.drawString(right_margin + x, top_margin + y, cell_value)

    # Sauvegarder le PDF
    c.save()  

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Imprimer la plage non vide d'un onglet Excel en PDF.")
    parser.add_argument("-f", "--file", required=True, help="Adresse du fichier Excel (.xlsx).")
    parser.add_argument("-w", "--worksheet", required=True, help="Nom de l'onglet à imprimer.")
    #parser.add_argument("-o", "--output", required=True, help="Sortie du fichier pdf")
    args = parser.parse_args()

    excel_file = args.file
    sheet_name = args.worksheet
    output_pdf = f"{sheet_name}.pdf"
    
    '''Création du fichier PDF sans avoir à mettre les paramètres en arguments
    excel_file = "E:\DOC Harinaivo\DataCenter_actuel.xlsx"
    sheet_name = "JLD2"
    output_pdf = "E:\DOC Harinaivo\JLD2.pdf"'''

    dimensions = get_excel_sheet_dimensions(excel_file, sheet_name)
    if dimensions:
        min_row, max_row, min_col, max_col = dimensions
        print(f"La feuille '{sheet_name}' du fichier '{excel_file}' a les dimensions suivantes:")
        print(f"Lignes : {min_row} à {max_row}")
        print(f"Colonnes : {min_col} à {max_col}")

    excel_sheet_to_pdf(excel_file, sheet_name, output_pdf)


 